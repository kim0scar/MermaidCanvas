#!/usr/bin/env python3
"""
Uppdaterar Visuali2e-v46-Funktionsprotokoll.xlsx med Status (sim) per funktion
baserat på STATUS-mappning från automated tests + explorativ validering.

Använder funktion-namnet (kolumn C) som nyckel för matchning.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

XLSX = Path("/Users/kim/2e Mermaid Code/Visuali2e-v46-Funktionsprotokoll.xlsx")

# ============================================================
# STATUS_MAP: funktionsnamn (matchar kolumn C i Excel) →
#             ("✓"|"✗"|"—", "kommentar/test-täckning")
#
# "✓" = PASS (verifierat fungerar i sim)
# "✗" = FAIL (känd bugg)
# "—" = EJ TESTAT (saknar coverage)
# ============================================================

STATUS_MAP = {
    # ===== Toolbar – Primärrad =====
    "Formväljare (chips)": ("✓", "DragOutTests + LayoutOverflowTests"),
    "Färgpaket-rad": ("—", "Ingen UI-test för färgapplicering"),
    "Textstil-rad": ("—", "Ingen UI-test för textstilar"),
    "Markeringsläge (multi-select)": ("—", "Ingen XCUITest för marker-mode"),
    "Zoom-badge": ("✓", "V35BugHuntTests testZoomChangesScrollViewState"),
    "Ångra (undo)": ("—", "Disabled vid start → ingen automated test"),
    "Lägen-meny (hamburger)": ("✓", "V33VersionVisibleTests + V29CoverageTests"),

    # ===== Hamburger-meny =====
    "Ny canvas (välj plattform)": ("—", "Ingen automated test"),
    "Aktuell plattform (indikator)": ("—", "Ingen automated test"),
    "Visa regler för Godot": ("—", "Plattform-specifik, ej testad"),
    "Spara": ("—", "FileSystem ej testat i UI"),
    "Spara som ny fil": ("—", "FileSystem ej testat"),
    "Öppna fil": ("—", "FileSystem ej testat"),
    "Importera Mermaid": ("—", "MermaidImportSheet ej UI-testad"),
    "Visa Mermaid-kod": ("✓", "V29CoverageTests testT22_ShowMermaidCodeMenuItemExists"),
    "Visa AppVersion": ("✓", "V33VersionVisibleTests testVersionIsVisibleInLägenMenu"),

    # ===== Form-chips =====
    "Cirkel": ("✓", "DragOutTests testTapAddsCircle + testDragCircleChipToCanvas"),
    "Rektangel": ("✓", "DragOutTests testDragRectangleChipToCanvas"),
    "Kvadrat": ("—", "Inget eget test, ingår i totalShapeCount"),
    "Romb (diamant)": ("✓", "DragOutTests testAllSixChipsProduceShapes"),
    "Piller (oval)": ("—", "Inget eget test, ingår i totalShapeCount"),
    "Processpil": ("—", "Inget eget test, ingår i totalShapeCount"),
    "Container (subgraph)": ("—", "Ingen UI-test för container-chip"),
    "Tabell": ("✓", "V27FeatureTests testTapTableChipAddsTable + testDragTableChipToCanvas"),
    "Jump-link (länkat par)": ("✓", "V27FeatureTests testTapLinkChipAddsJumpLinkPair"),
    "Linje (FreeLine)": ("—", "Ingen UI-test för line-chip"),
    "Antecknings-popup": ("—", "Ingen UI-test"),

    # ===== Form-paket =====
    "UI-paket": ("—", "Bara packs-toggle existens testad"),
    "Roadmap-paket": ("—", "Ingen test"),
    "Arkitektur-paket": ("✓", "V29CoverageTests testT21_ArchitecturePackTogglesAddsChipInShapesRow"),
    "Flow-paket": ("—", "Ingen test"),

    # ===== Färg-rad =====
    "Applicera färgpaket": ("—", "Round-trip täcker JSON, ej UI-flödet"),
    "Återställ till kategori-färg": ("—", "Ingen UI-test"),

    # ===== Textstil-rad =====
    "Textstorlek-popup": ("—", "Ingen UI-test"),
    "Fet text-toggle": ("—", "Ingen UI-test"),
    "Punktlista": ("✓ (data)", "Round-trip via V35MermaidValidationTests testRoundTrip_v46Fields_Preserved"),
    "Numrerad lista": ("✓ (data)", "Round-trip via testRoundTrip_v46Fields_Preserved"),
    "Justering – vänster": ("—", "UI-knapparna i textStylesSecondary ej UI-testade"),
    "Justering – centrera": ("—", "Ingen UI-test"),
    "Justering – höger": ("—", "Ingen UI-test"),
    "Indrag minska": ("✓ (data)", "Round-trip via testRoundTrip_v46Fields_Preserved"),
    "Indrag öka": ("✓ (data)", "Round-trip via testRoundTrip_v46Fields_Preserved"),

    # ===== Multi-select rad =====
    "Räknare (N markerade)": ("—", "Ingen test"),
    "Duplicera markerade": ("—", "Ingen test"),
    "Ta bort markerade": ("—", "Ingen test"),
    "Centrera horisontellt": ("—", "Ingen test"),
    "Centrera vertikalt": ("—", "Ingen test"),

    # ===== Canvas-gester =====
    "Tap på form → välj": ("—", "Implicit täckt via testT17_DoubleTapOpensEdit"),
    "Tap på form i marker-läge": ("—", "Ingen test"),
    "Drag form → flytta": ("✓", "V33SensorTests testDragCircleToKnownPosition"),
    "Drag i multi-select → flytta alla": ("—", "Ingen test"),
    "Marquee-drag → multi-select": ("—", "Ingen test"),
    "Tap på tom canvas → avmarkera": ("✓", "V29CoverageTests testT16_TapBackgroundDeselects"),
    "Long-press på form → ConnectionHandle": ("—", "Ingen test"),
    "Pinch zoom": ("✓", "V35BugHuntTests testZoomChangesScrollViewState"),
    "Pan canvas": ("✓", "V34PanSymmetryTests testPanWorksInAllFourDirections"),

    # ===== Selection handles =====
    "Proportionell resize": ("—", "Ingen test"),
    "Fri resize": ("—", "Ingen test"),
    "Rotation": ("—", "Ingen test"),
    "Multi-select resize": ("—", "Ingen test"),
    "Connection handle drag": ("—", "Ingen test (komplex gesture)"),

    # ===== Form-actions =====
    "Dubbelklick → öppna editor": ("✓", "V29CoverageTests testT17_DoubleTapOpensEdit"),
    "Tap på note-badge → mini-sheet": ("—", "Ingen test"),
    "Collapse-badge på container": ("—", "Ingen test"),

    # ===== EditShapeSheet =====
    "Redigera text (label)": ("✓", "EndToEndTests testNoteTypedInEditSheetFollowsToMermaidCode"),
    "Toggla 'Visa text'": ("—", "Ingen UI-test"),
    "Stilväljare (segmented)": ("—", "Round-trip via testRoundTrip_PreservesAllShapeMetadata"),
    "Justering + bullets (inline)": ("✓ (data)", "Via testRoundTrip_v46Fields_Preserved"),
    "Redigera anteckning (note)": ("✓", "EndToEndTests testNoteTypedInEditSheetFollowsToMermaidCode"),
    "Ta bort form (via sheet)": ("—", "Ingen UI-test"),

    # ===== Kant-context =====
    "Tap på pil → context-meny": ("—", "Ingen test"),
    "Lägg till/ändra etikett": ("✓ (data)", "Round-trip via testRoundTrip_PreservesEdgeLabelsAndStyles_FullCycle"),
    "Ändra pil-riktning": ("✓ (data)", "Round-trip via testRoundTrip_EdgesWithStyles"),
    "Ändra linje-stil": ("✓ (data)", "Round-trip via testRoundTrip_EdgesWithStyles"),
    "Ta bort kant": ("—", "Ingen UI-test"),

    # ===== Tabell-editor =====
    "Öppna tabell-editor": ("—", "Ingen UI-test"),
    "Redigera tabell-cell": ("✓ (data)", "Round-trip via testRoundTrip_v46Fields_Preserved (tableCells)"),
    "Lägg till rad": ("—", "Ingen UI-test"),
    "Lägg till kolumn": ("—", "Ingen UI-test"),

    # ===== Ny canvas-sheet =====
    "Välj 'Blank canvas'": ("—", "Ingen UI-test"),
    "Välj 'Godot'": ("—", "Ingen UI-test"),

    # ===== Note-sheets =====
    "NoteMiniSheet (per form)": ("—", "Ingen UI-test"),
    "NotePopupSheet (alla)": ("—", "Ingen UI-test"),

    # ===== Mermaid-export =====
    "Visa Mermaid-kod": ("✓", "V29CoverageTests testT22_ShowMermaidCodeMenuItemExists"),
    "Kopiera kod till urklipp": ("—", "Ingen UI-test"),

    # ===== Mermaid-import =====
    "Steg 1: Visa mall för AI": ("—", "Ingen UI-test"),
    "Steg 1: Kopiera mall": ("—", "Ingen UI-test"),
    "Steg 2: Klistra in Mermaid-kod": ("—", "Ingen UI-test"),
    "Steg 2: Importera": ("—", "Ingen UI-test"),
    "Navigera mellan steg": ("—", "Ingen UI-test"),

    # ===== Plattform-regler =====
    "Visa plattform-regler": ("—", "Ingen UI-test"),

    # ===== Färgväljare =====
    "ColorPickerPopover – välj färg": ("—", "Ingen UI-test"),
    "ColorPickerPopover – återställ override": ("—", "Ingen UI-test"),
    "ColorPackPopover – välj paket": ("✓ (data)", "Round-trip via testGenerator_ColorPack"),

    # ===== Canvas auto =====
    "Automatisk canvas-expansion": ("✓", "V27FeatureTests testCircleLandsNearCanvasEdge"),
    "Auto-scroll under drag": ("—", "Svår att testa automated"),
    "Live update vid externa ändringar": ("—", "NSFilePresenter ej testad i sim"),

    # ===== FreeLine =====
    "Drag linje-slutpunkt": ("✓ (data)", "Round-trip via testRoundTrip_LineEnd_Preserved"),

    # ===== Container =====
    "Dra form in i container": ("—", "Ingen UI-test"),
    "Dra container → barn följer med": ("—", "Ingen UI-test"),
    "Resize container → barn skalas inte": ("—", "Ingen UI-test"),

    # ===== Canvas snap =====
    "Dot-grid bakgrund": ("✓", "Visuell verifiering i launch-screenshot"),

    # ===== Persistens =====
    "Autospara vid bakgrundning": ("—", "Ingen test"),
    "Round-trip av alla form-fält": ("✓", "FULLT TÄCKT: V35MermaidValidationTests + RoundTripTests"),
}


# ============================================================
# Funktioner som är BUGGAR / FINDINGS från valideringen
# ============================================================
FINDINGS = {
    # "Funktion-namn": ("✗", "Bugg-beskrivning")
    # Lägg till här när vi hittar konkreta buggar
}


def update_xlsx():
    wb = load_workbook(XLSX)
    ws = wb.active

    # Hitta header-rad (ska vara rad 4)
    header_row = 4
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]

    func_col = headers.index("Funktion") + 1
    status_sim_col = headers.index("Status (simulator)") + 1
    note_col = headers.index("Anteckning / bugg-ID") + 1

    matched = 0
    unmatched_names = []
    finding_count = 0

    for row in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row, column=func_col).value
        if not name:
            continue

        # Kolla först FINDINGS (buggar)
        if name in FINDINGS:
            status, note = FINDINGS[name]
            ws.cell(row=row, column=status_sim_col).value = status
            ws.cell(row=row, column=note_col).value = note
            ws.cell(row=row, column=status_sim_col).font = Font(name="Helvetica", size=10, bold=True, color="C00000")
            finding_count += 1
            matched += 1
            continue

        # Sedan vanlig STATUS_MAP
        if name in STATUS_MAP:
            status, note = STATUS_MAP[name]
            ws.cell(row=row, column=status_sim_col).value = status
            ws.cell(row=row, column=note_col).value = note

            color = {
                "✓": "1A7A3C",
                "✓ (data)": "8A9A2E",  # gulgrön - bara data-lagret
                "✗": "C00000",
                "—": "888888",
            }.get(status, "000000")
            ws.cell(row=row, column=status_sim_col).font = Font(name="Helvetica", size=11, bold=True, color=color)
            ws.cell(row=row, column=status_sim_col).alignment = Alignment(horizontal="center", vertical="center")
            matched += 1
        else:
            unmatched_names.append(name)

    wb.save(XLSX)

    print(f"✓ Uppdaterade {matched} rader")
    print(f"  - varav {finding_count} markerade som buggar")
    print(f"✗ {len(unmatched_names)} omappade funktionsnamn:")
    for n in unmatched_names[:20]:
        print(f"    - {n}")

    # Sammanfattning av status-fördelning
    counts = {"✓": 0, "✓ (data)": 0, "✗": 0, "—": 0}
    for status, _ in STATUS_MAP.values():
        counts[status] = counts.get(status, 0) + 1
    print(f"\nStatus-fördelning:")
    print(f"  ✓ PASS (full):       {counts['✓']}")
    print(f"  ✓ PASS (data-only):  {counts['✓ (data)']}")
    print(f"  ✗ FAIL:              {counts['✗']}")
    print(f"  — EJ TESTAT:          {counts['—']}")
    print(f"  Total:                {sum(counts.values())}")


if __name__ == "__main__":
    update_xlsx()
