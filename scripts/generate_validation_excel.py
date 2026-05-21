#!/usr/bin/env python3
"""
Genererar Visuali2e-v46-Funktionsprotokoll.xlsx — uttömmande funktionsöversikt
+ valideringsprotokoll för Kim.

Ett blad, kolumner:
  Nr | Kategori | Funktion | Beskrivning | Åtkomst | Förväntad effekt | Status (sim) | Status (iPhone) | Anteckning
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


# (Kategori, Funktion, Beskrivning, Åtkomst, Förväntad effekt)
FUNCTIONS = [
    # ============== Toolbar – Primär rad ==============
    ("Toolbar – Primärrad", "Formväljare (chips)",
     "Öppnar två rader med alla 12 former (cirkel, rektangel, kvadrat, romb, piller, process-pil, container, tabell, jump-link, linje, antecknings-popup).",
     "Tap på square.on.circle-ikonen i toolbar.",
     "Sekundär rad expanderar med shape-chips. Tap igen stänger raden."),

    ("Toolbar – Primärrad", "Färgpaket-rad",
     "Öppnar rad med 7 färgpaket (UI, Roadmap, Arkitektur, Flow + ingen-färg) för aktuell vald form.",
     "Tap på paintpalette-ikonen (disabled om ingen form vald).",
     "Sekundär rad visar färgcirklar; tap applicerar paket på vald form."),

    ("Toolbar – Primärrad", "Textstil-rad",
     "Öppnar rad med textstorlek-val, fet, punktlista, numrerad lista, justering L/C/R, indrag in/ut.",
     "Tap på textformat.size-ikonen (disabled om ingen form vald).",
     "Sekundär rad expanderar med text-operationsknappar."),

    ("Toolbar – Primärrad", "Markeringsläge (multi-select)",
     "Aktivera marquee-läge för att rita streckad rektangel och välja flera former i ett svep.",
     "Tap på rectangle.dashed-ikonen.",
     "MarkerOverlay aktiveras; markeringsspecifik toolbar-rad visas; drag på canvas ritar streckad rektangel."),

    ("Toolbar – Primärrad", "Zoom-badge",
     "Visar aktuell zoom-procent och fungerar som debug-räknare (shapeCount=N). Tap återställer zoom till 100%.",
     "Tap på toolbar.zoom (procent-badgen).",
     "Zoom återställs till 1.0; canvas centreras på nuvarande viewport."),

    ("Toolbar – Primärrad", "Ångra (undo)",
     "Ångrar senaste åtgärd (skapa, flytta, ta bort, ändra text, etc).",
     "Tap på arrow.uturn.backward-ikonen.",
     "Föregående snapshot återställs. Disabled när inget finns att ångra."),

    ("Toolbar – Primärrad", "Lägen-meny (hamburger)",
     "Öppnar dropdown-meny med fil-operationer (Ny, Spara, Öppna), plattform, Mermaid-kod.",
     "Tap på slider.horizontal.3-ikonen längst till höger.",
     "Dropdown-meny visas med menyval."),

    # ============== Hamburger-meny (Lägen) ==============
    ("Hamburger-meny", "Ny canvas (välj plattform)",
     "Startar en helt ny rityta. Användaren får välja plattform (Blank eller Godot).",
     "Tap hamburger → 'Ny canvas (välj plattform)' (allra översta valet).",
     "NewCanvasSheet öppnas; vald plattform skapar tom canvas."),

    ("Hamburger-meny", "Aktuell plattform (indikator)",
     "Visar vilken plattform aktuell canvas använder. Informativ rad – ej klickbar.",
     "Visas direkt under 'Ny canvas' som disabled grå text.",
     "Text: 'Aktuell plattform: Blank canvas' eller 'Aktuell plattform: Godot'."),

    ("Hamburger-meny", "Visa regler för Godot",
     "Öppnar regelbok med plattform-specifika regler (visas bara om Godot-plattform är vald).",
     "Tap hamburger → 'Visa regler för Godot' (visas endast i Godot-läge).",
     "PlatformRulesSheet öppnas med markdown-formaterad regeltext."),

    ("Hamburger-meny", "Spara",
     "Sparar canvas till senast öppnad fil (JSON inbäddat i Mermaid-MD).",
     "Tap hamburger → 'Spara'. Heter 'Spara…' om ingen fil är öppen.",
     "Fil skrivs till disk; status-text bekräftar."),

    ("Hamburger-meny", "Spara som ny fil",
     "Sparar canvas till en ny fil (filväljare).",
     "Tap hamburger → 'Spara som ny fil…'.",
     "Filväljare öppnas för att välja sparplats; ny fil skapas."),

    ("Hamburger-meny", "Öppna fil",
     "Öppnar tidigare sparad Mermaid-MD-fil och läser in canvas-state.",
     "Tap hamburger → 'Öppna fil…'.",
     "Filväljare öppnas; vald fil parsas och canvas ersätts."),

    ("Hamburger-meny", "Importera Mermaid",
     "Två-stegs guide för att importera AI-genererad Mermaid-kod från Claude.ai eller annan AI.",
     "Tap hamburger → 'Importera Mermaid…'.",
     "MermaidImportSheet öppnas (steg 1: kopiera mall, steg 2: klistra in)."),

    ("Hamburger-meny", "Visa Mermaid-kod",
     "Visar den live-genererade Mermaid-koden för canvas. Kan kopieras och klistras in i Claude eller mermaid.live.",
     "Tap hamburger → 'Visa Mermaid-kod'.",
     "MermaidCodeSheet öppnas med kod-text och 'Kopiera'-knapp."),

    ("Hamburger-meny", "Visa AppVersion",
     "Visar appens versionsnummer som info-rad i menyn.",
     "Disabled rad längst ner i hamburger-menyn.",
     "Visar 'v46' (eller aktuell version)."),

    # ============== Forms toolbar — Chips (rad A + B) ==============
    ("Form-chips", "Cirkel",
     "Lägg till en cirkelform på canvas.",
     "Tap toolbar.shapes → tap chip.circle (eller långpress + drag till canvas-position).",
     "Cirkelform skapas vid canvas-mitten eller släpp-punkten."),

    ("Form-chips", "Rektangel",
     "Lägg till en rektangelform på canvas.",
     "Tap toolbar.shapes → tap chip.rectangle (eller drag).",
     "Rektangelform skapas."),

    ("Form-chips", "Kvadrat",
     "Lägg till en kvadratisk form (80×80) på canvas.",
     "Tap toolbar.shapes → tap chip.square (eller drag).",
     "Kvadrat skapas."),

    ("Form-chips", "Romb (diamant)",
     "Lägg till en romb/diamantform på canvas.",
     "Tap toolbar.shapes → tap chip.diamond (eller drag).",
     "Rombform med rundade hörn skapas."),

    ("Form-chips", "Piller (oval)",
     "Lägg till en avlång ovalform på canvas.",
     "Tap toolbar.shapes → tap chip.pill (eller drag).",
     "Pillerformad oval skapas."),

    ("Form-chips", "Processpil",
     "Lägg till en pentagon-formad processpil på canvas.",
     "Tap toolbar.shapes → tap chip.processArrow (eller drag).",
     "Pentagon med spets åt höger skapas."),

    ("Form-chips", "Container (subgraph)",
     "Lägg till en streckad behållare för att gruppera andra former.",
     "Tap toolbar.shapes → tap chip.container (eller drag).",
     "Större streckad rektangel skapas; andra former kan dras in i den."),

    ("Form-chips", "Tabell",
     "Lägg till en 3×3-tabellform som kan redigeras cell för cell.",
     "Tap toolbar.shapes → tap chip.table.",
     "Tabellform skapas; dubbelklick öppnar TableEditorSheet."),

    ("Form-chips", "Jump-link (länkat par)",
     "Lägg till två länkade former (för att hoppa mellan ställen på canvas).",
     "Tap toolbar.shapes → tap chip.link.",
     "Två länk-former med samma nummer skapas nära canvas-mitten."),

    ("Form-chips", "Linje (FreeLine)",
     "Lägg till en lös linje (ingen pil) som kan dras i båda ändar.",
     "Tap toolbar.shapes → tap chip.line.",
     "Horisontell linje skapas; båda ändpunkter draggable."),

    ("Form-chips", "Antecknings-popup",
     "Öppnar list-vy med alla textfält (labels + notes) på canvas.",
     "Tap toolbar.shapes → tap chip.notepopup (pratbubbla-ikon).",
     "NotePopupSheet öppnas med scrollable lista."),

    # ============== Form-paket (toolbar) ==============
    ("Form-paket", "UI-paket",
     "Aktivera/avaktivera UI-form-chips i toolbaren (designsystem-ikoner).",
     "Tap toolbar.packs → toggle 'UI'.",
     "UI-pack-chips visas/döljs i shapes-raden."),

    ("Form-paket", "Roadmap-paket",
     "Aktivera/avaktivera Roadmap-form-chips (milstolpar, deadlines, etc).",
     "Tap toolbar.packs → toggle 'Roadmap'.",
     "Roadmap-pack-chips visas/döljs."),

    ("Form-paket", "Arkitektur-paket",
     "Aktivera/avaktivera Arkitektur-form-chips (komponenter, databas, etc).",
     "Tap toolbar.packs → toggle 'Arkitektur'.",
     "Arkitektur-pack-chips visas/döljs."),

    ("Form-paket", "Flow-paket",
     "Aktivera/avaktivera Flow-form-chips (extra flödesschema-symboler).",
     "Tap toolbar.packs → toggle 'Flow'.",
     "Flow-pack-chips visas/döljs."),

    # ============== Färg-toolbar ==============
    ("Färg-rad", "Applicera färgpaket",
     "Sätt en färgkombination (fill+stroke+text) på vald form.",
     "Välj form först → tap toolbar.colors → tap färg-cirkel.",
     "Vald forms colorPackId uppdateras; visuellt utseende ändras."),

    ("Färg-rad", "Återställ till kategori-färg",
     "Ta bort form-specifik färg-override; använd kategori-standardfärg.",
     "Tap 'ingen färg'-cirkel (markerad med snedstreck).",
     "colorPackId blir nil; formen använder kategori-färg."),

    # ============== Textstilar-rad ==============
    ("Textstil-rad", "Textstorlek-popup",
     "Välj mellan Rubrik 1, Rubrik 2, Rubrik 3 eller Brödtext.",
     "Tap toolbar.textstyles → tap textformat.size → välj i confirmationDialog.",
     "Vald forms textStyle ändras; text storlek/vikt uppdateras."),

    ("Textstil-rad", "Fet text-toggle",
     "Togglar mellan Rubrik 1 (fet) och Brödtext (normal vikt).",
     "Tap toolbar.textstyles → tap 'bold'-knappen.",
     "Text byter mellan headline (.r1) och body."),

    ("Textstil-rad", "Punktlista",
     "Lägger till • framför varje rad i texten. Stänger av numrerad lista.",
     "Tap toolbar.textstyles → tap 'list.bullet'.",
     "Shape.hasBullets togglas; hasNumberedList=false; text re-renderas."),

    ("Textstil-rad", "Numrerad lista",
     "Lägger till 1./2./3. framför varje rad. Stänger av punktlista.",
     "Tap toolbar.textstyles → tap 'list.number'.",
     "Shape.hasNumberedList togglas; hasBullets=false."),

    ("Textstil-rad", "Justering – vänster",
     "Vänsterjustera texten i formen.",
     "Tap toolbar.textstyles → tap 'text.alignleft'.",
     "Shape.textAlignment = .leading; text re-renderas."),

    ("Textstil-rad", "Justering – centrera",
     "Centrera texten i formen.",
     "Tap toolbar.textstyles → tap 'text.aligncenter'.",
     "Shape.textAlignment = .center."),

    ("Textstil-rad", "Justering – höger",
     "Högerjustera texten i formen.",
     "Tap toolbar.textstyles → tap 'text.alignright'.",
     "Shape.textAlignment = .trailing."),

    ("Textstil-rad", "Indrag minska",
     "Minskar indragsnivån (0–3 steg).",
     "Tap toolbar.textstyles → tap 'decrease.indent'.",
     "Shape.indentLevel -= 1 (min 0); text indenteras mindre."),

    ("Textstil-rad", "Indrag öka",
     "Ökar indragsnivån (0–3 steg).",
     "Tap toolbar.textstyles → tap 'increase.indent'.",
     "Shape.indentLevel += 1 (max 3); text indenteras mer."),

    # ============== MultiSelect-toolbar ==============
    ("Multi-select rad", "Räknare (N markerade)",
     "Visar antal valda former i marker-läge. Informativ rad.",
     "Aktivera marker-läge + välj former → räknare uppdateras.",
     "'1 markerad' / '5 markerade' visas."),

    ("Multi-select rad", "Duplicera markerade",
     "Skapar kopior av alla markerade former.",
     "Marker-läge + minst 1 form vald → tap 'plus.square.on.square'.",
     "Alla valda former dupliceras med offset; multiSelection uppdateras till nya."),

    ("Multi-select rad", "Ta bort markerade",
     "Raderar alla markerade former och deras anslutna pilar.",
     "Marker-läge + minst 1 form vald → tap 'trash' (röd).",
     "Alla valda former + tillhörande kanter tas bort."),

    ("Multi-select rad", "Centrera horisontellt",
     "Justerar alla markerade former så de delar samma X-center.",
     "Marker-läge + minst 2 former → tap 'align.horizontal.center'.",
     "Alla former får samma X-position; behåller sina Y-positioner."),

    ("Multi-select rad", "Centrera vertikalt",
     "Justerar alla markerade former så de delar samma Y-center.",
     "Marker-läge + minst 2 former → tap 'align.vertical.center'.",
     "Alla former får samma Y-position; behåller sina X-positioner."),

    # ============== Canvas-gester ==============
    ("Canvas-gester", "Tap på form → välj",
     "Markera en enskild form för att redigera den.",
     "Tap på valfri form på canvas.",
     "Formen får blå streckad markerings-ram; selection handles visas."),

    ("Canvas-gester", "Tap på form i marker-läge",
     "Lägg till/ta bort form från multi-selection.",
     "Aktivera marker-läge → tap på form.",
     "Formen togglas in/ur multiSelection; bounding box uppdateras."),

    ("Canvas-gester", "Drag form → flytta",
     "Flytta en form genom att dra den med fingret.",
     "Long-press (eller direkt drag) på form → drag till ny position.",
     "Form följer fingret; auto-scroll vid kant; canvas expanderas vid behov."),

    ("Canvas-gester", "Drag i multi-select → flytta alla",
     "Flytta alla markerade former samtidigt.",
     "Multi-select aktiv + drag på en vald form.",
     "Alla former i multiSelection (samt container-barn) följer med."),

    ("Canvas-gester", "Marquee-drag → multi-select",
     "Rita streckad rektangel för att markera flera former samtidigt.",
     "Aktivera marker-läge → drag på tom canvas-yta (>8pt).",
     "Streckad rektangel ritas; alla former inom blir multi-selected."),

    ("Canvas-gester", "Tap på tom canvas → avmarkera",
     "Töm selection när man tappar på tom yta.",
     "Tap på tom yta utanför former.",
     "selectedShapeId = nil; multiSelection töms; edge-mode avbryts om aktivt."),

    ("Canvas-gester", "Long-press på form → ConnectionHandle",
     "Visa pil-handtag på formens högra sida för att skapa kant.",
     "Långt tryck på form (≈0.5s).",
     "Blå ConnectionHandle dyker upp; drag därifrån skapar pil."),

    ("Canvas-gester", "Pinch zoom",
     "Zooma in/ut canvas med två-fingers nyp.",
     "Pinch-gest på canvas (två fingrar).",
     "Zoom-nivå uppdateras; pinch-anchor stannar där fingrarna är (UIScrollView)."),

    ("Canvas-gester", "Pan canvas",
     "Panorera canvas (förflytta viewport).",
     "Drag med två fingrar (eller en finger på tom yta utanför form).",
     "Viewport-offset uppdateras inom canvas-gränser."),

    # ============== Selection handles ==============
    ("Selection handles", "Proportionell resize",
     "Skala vald form proportionellt (aspect ratio bevaras).",
     "Drag på proportional-handtag (bottom-right hörn).",
     "Shape växer/krymper proportionellt från sitt center."),

    ("Selection handles", "Fri resize",
     "Skala vald form fritt i båda dimensioner (kan ändra aspect).",
     "Drag på free-resize-handtag (bottom-left, 4-pil-ikon).",
     "Width och height ändras oberoende."),

    ("Selection handles", "Rotation",
     "Rotera vald form kring sitt center.",
     "Drag på rotate-handtag (top-left).",
     "Shape.rotation uppdateras (grader)."),

    ("Selection handles", "Multi-select resize",
     "Skala alla markerade former proportionellt från gruppens center.",
     "Multi-select aktiv → drag på multi-resize-handtag (bottom-right av bounding box).",
     "Alla markerade former skalas proportionellt; relativ position bevaras."),

    ("Selection handles", "Connection handle drag",
     "Skapa pil mellan två former via gummibands-drag.",
     "Long-press på form → drag från ConnectionHandle till en annan form.",
     "Pil skapas mellan from- och to-form; default direction=forward."),

    # ============== Form-context (lokala menyer) ==============
    ("Form-actions", "Dubbelklick → öppna editor",
     "Öppnar EditShapeSheet för vald form (text, stil, anteckning).",
     "Dubbelklick på form (eller tap på 'Edit' i context-meny om sådan finns).",
     "EditShapeSheet öppnas med formens nuvarande värden."),

    ("Form-actions", "Tap på note-badge → mini-sheet",
     "Öppnar liten popup för att läsa/redigera formens anteckning.",
     "Tap på gul note-badge (om note finns) på formen.",
     "NoteMiniSheet öppnas över formen."),

    ("Form-actions", "Collapse-badge på container",
     "Visar att container har dolda barn; tap visar dem igen.",
     "Tap på collapse-badge under container-form.",
     "Container-barn visas/döljs."),

    # ============== EditShapeSheet ==============
    ("EditShapeSheet", "Redigera text (label)",
     "Skriv eller ändra texten som visas i formen.",
     "Dubbelklick på form → TextField 'Skriv text' överst.",
     "Shape.label uppdateras live på canvas vid Klar/Spara."),

    ("EditShapeSheet", "Toggla 'Visa text'",
     "Visa/dölj formens label på canvas (sparas i state).",
     "EditShapeSheet → Toggle 'Visa text'.",
     "Shape.showLabel togglas; text visas/försvinner på formen."),

    ("EditShapeSheet", "Stilväljare (segmented)",
     "Välj textstil (Rubrik 1/2/3 eller Brödtext) via segmented picker.",
     "EditShapeSheet → Picker 'Stil'.",
     "Shape.textStyle uppdateras."),

    ("EditShapeSheet", "Justering + bullets (inline)",
     "Snabb-rad med L/C/R-justering och bullet-toggle.",
     "EditShapeSheet → HStack under stilväljare.",
     "Shape.textAlignment och hasBullets uppdateras."),

    ("EditShapeSheet", "Redigera anteckning (note)",
     "Skriv en privat anteckning som inte syns direkt på canvas.",
     "EditShapeSheet → TextEditor 'Skriv anteckning här'.",
     "Shape.note uppdateras; note-badge visas på formen."),

    ("EditShapeSheet", "Ta bort form (via sheet)",
     "Radera form från canvas via en knapp i sheet:t.",
     "EditShapeSheet → 'Ta bort form'-knapp → BekräftelseDialog → Bekräfta.",
     "Formen och dess pilar tas bort; sheet stängs."),

    # ============== Kant-operationer ==============
    ("Kant-context", "Tap på pil → context-meny",
     "Visar context-meny för pil (label, direction, style, delete).",
     "Tap på pil/kant på canvas.",
     "Context-meny visas vid pilen."),

    ("Kant-context", "Lägg till/ändra etikett",
     "Skriv text på en pil.",
     "Tap pil → context-meny → 'Edit label' (eller dubbelklick på pil).",
     "EdgeLabelSheet öppnas; text sparas på kanten."),

    ("Kant-context", "Ändra pil-riktning",
     "Välj fram-pil (→), bak-pil (←), dubbelriktad (↔) eller ingen (—).",
     "Tap pil → midpoint-knapp eller context-meny → välj direction.",
     "EdgeConnection.direction uppdateras; pilhuvud ritas om."),

    ("Kant-context", "Ändra linje-stil",
     "Växla mellan solid och streckad linje.",
     "Tap pil → context-meny → byt style (solid/dashed).",
     "EdgeConnection.style uppdateras."),

    ("Kant-context", "Ta bort kant",
     "Radera pilen mellan två former.",
     "Tap pil → context-meny → 'Delete' (eller 'Ta bort').",
     "EdgeConnection raderas från modellen."),

    # ============== TableEditorSheet ==============
    ("Tabell-editor", "Öppna tabell-editor",
     "Öppnar redigerings-sheet för tabell-form.",
     "Dubbelklick på tabell-form på canvas.",
     "TableEditorSheet öppnas med rutnät av celler."),

    ("Tabell-editor", "Redigera tabell-cell",
     "Skriv text i en enskild cell.",
     "Tap i cell i TableEditorSheet → skriv text.",
     "Cell uppdateras; sparas när sheet:t stängs."),

    ("Tabell-editor", "Lägg till rad",
     "Lägg till en ny tom rad längst ner i tabellen.",
     "TableEditorSheet → 'Lägg till rad'-knapp.",
     "Ny rad läggs till; rutnätet växer en rad."),

    ("Tabell-editor", "Lägg till kolumn",
     "Lägg till en ny tom kolumn till höger i tabellen.",
     "TableEditorSheet → 'Lägg till kolumn'-knapp.",
     "Ny kolumn läggs till; rutnätet växer en kolumn."),

    # ============== NewCanvasSheet ==============
    ("Ny canvas-sheet", "Välj 'Blank canvas'",
     "Starta tom canvas utan plattform-regler.",
     "Hamburger → Ny canvas → tap 'Blank canvas'-kortet.",
     "Tom canvas skapas; platform = .blank."),

    ("Ny canvas-sheet", "Välj 'Godot'",
     "Starta canvas med Godot-plattformsregler (iPhone-ram-overlay etc).",
     "Hamburger → Ny canvas → tap 'Godot'-kortet.",
     "Canvas skapas; platform = .godot; specifika regler aktiveras."),

    # ============== Antecknings-sheets ==============
    ("Note-sheets", "NoteMiniSheet (per form)",
     "Liten popup för att läsa/skriva anteckning på en form.",
     "Tap note-badge på form, eller via EditShapeSheet.",
     "TextEditor visas; ändringar sparas till shape.note."),

    ("Note-sheets", "NotePopupSheet (alla)",
     "Sammanställning av alla labels + notes på canvas i en lista.",
     "Tap chip.notepopup (pratbubbla-ikonen i shapes-raden).",
     "ScrollView med alla former och deras text visas."),

    # ============== MermaidCodeSheet (export) ==============
    ("Mermaid-export", "Visa Mermaid-kod",
     "Visar live-genererad Mermaid-syntax för aktuell canvas.",
     "Hamburger → 'Visa Mermaid-kod'.",
     "MermaidCodeSheet öppnas; text uppdateras vid varje canvas-ändring."),

    ("Mermaid-export", "Kopiera kod till urklipp",
     "Kopiera Mermaid-koden för att klistra in i Claude.ai eller mermaid.live.",
     "MermaidCodeSheet → 'Kopiera'-knapp.",
     "Kod hamnar i urklipp; knappen visar 'Kopierad' i 1.5s."),

    # ============== MermaidImportSheet ==============
    ("Mermaid-import", "Steg 1: Visa mall för AI",
     "Visa instruktionstext att skicka till Claude.ai så den genererar rätt Mermaid-syntax.",
     "Hamburger → 'Importera Mermaid' → Steg 1.",
     "Mall-text visas; kan kopieras."),

    ("Mermaid-import", "Steg 1: Kopiera mall",
     "Kopierar instruktionstexten till urklipp.",
     "Steg 1 → 'Kopiera mall'-knapp.",
     "Mall hamnar i urklipp."),

    ("Mermaid-import", "Steg 2: Klistra in Mermaid-kod",
     "Klistra in AI-genererad Mermaid-kod i textfält.",
     "Steg 2 → tap i TextEditor → paste från urklipp.",
     "Kod sparas i state; preview kan visas."),

    ("Mermaid-import", "Steg 2: Importera",
     "Parsa Mermaid-koden och ersätt canvas med resultatet.",
     "Steg 2 → 'Importera till canvas'-knapp.",
     "CanvasModel.replaceAll() körs; canvas uppdateras; sheet stängs."),

    ("Mermaid-import", "Navigera mellan steg",
     "Gå mellan steg 1 (mall) och steg 2 (klistra in).",
     "Knappar 'Nästa' / 'Tillbaka' i sheet:t.",
     "Steg-vy byts; värden bevaras."),

    # ============== PlatformRulesSheet ==============
    ("Plattform-regler", "Visa plattform-regler",
     "Visar markdown-formaterad regelbok för aktuell plattform (t.ex. Godot).",
     "Hamburger → 'Visa regler för [plattform]'.",
     "PlatformRulesSheet öppnas med ScrollView."),

    # ============== Färgväljare (popovers) ==============
    ("Färgväljare", "ColorPickerPopover – välj färg",
     "Välj en specifik färg-override för en form (8 fördefinierade).",
     "Långpress eller meny-val på form → ColorPickerPopover.",
     "Grid med 8 färger visas; vald sätter colorOverride."),

    ("Färgväljare", "ColorPickerPopover – återställ override",
     "Ta bort form-specifik färg så att kategori-paket används.",
     "ColorPickerPopover → 'Använd kategori-färg'-knapp.",
     "colorOverride blir nil; formen visar paketets färg."),

    ("Färgväljare", "ColorPackPopover – välj paket",
     "Välj färgpaket (7 paket: UI, Roadmap, Arkitektur, Flow + 2 + ingen).",
     "Toolbar.colors → tap färgcirkel.",
     "colorPackId uppdateras; form-utseende ändras direkt."),

    # ============== Canvas-generella beteenden ==============
    ("Canvas – auto", "Automatisk canvas-expansion",
     "Canvas växer automatiskt när form dras nära kanten.",
     "Drag form mot canvas-kant (inom 50pt marginal).",
     "Canvas-storlek växer; viewport scrollar med."),

    ("Canvas – auto", "Auto-scroll under drag",
     "Viewport panorerar långsamt när form dras mot viewport-kant.",
     "Drag form mot synligt kant-område.",
     "Viewport glider långsamt för att följa formen."),

    ("Canvas – auto", "Live update vid externa ändringar",
     "Om någon annan ändrar canvas-filen i iCloud uppdateras vyn live.",
     "Spara extern Mermaid-MD i iCloud-mappen.",
     "Canvas läses om automatiskt (NSFilePresenter)."),

    # ============== Linje-/pil-specifika ==============
    ("FreeLine-specifik", "Drag linje-slutpunkt",
     "Förläng/förkorta/rotera en fri linje genom att dra dess ände.",
     "Tap linje → drag på endpoint-handtag.",
     "Shape.lineEnd uppdateras; linje följer."),

    # ============== Container-specifika ==============
    ("Container", "Dra form in i container",
     "När en form släpps inom en container blir den container-barn.",
     "Drag form till position inom container-rektangel.",
     "Form blir barn till container; följer med när container flyttas."),

    ("Container", "Dra container → barn följer med",
     "När container flyttas följer alla dess barn-former med automatiskt.",
     "Drag på container-form.",
     "Container + alla barn flyttas tillsammans."),

    ("Container", "Resize container → barn skalas inte",
     "När container skalas påverkas inte barn-formers storlek.",
     "Drag på containerns resize-handtag.",
     "Endast containerns dimensioner ändras; barn behåller sina storlekar."),

    # ============== Snap & alignment (om relevant) ==============
    ("Canvas – snap", "Dot-grid bakgrund",
     "Visar prick-rutnät som visuell guide.",
     "Alltid synlig som bakgrund (DotGridBackground).",
     "Prickar med ~40pt mellanrum; rörs ej av zoom."),

    # ============== Persistence ==============
    ("Persistens", "Autospara vid bakgrundning",
     "Canvas sparas automatiskt när appen läggs i bakgrunden.",
     "Tryck hem-knappen / app switch.",
     "Aktuell fil skrivs till disk (om en är öppen)."),

    ("Persistens", "Round-trip av alla form-fält",
     "Alla form-egenskaper bevaras vid spara → öppna (numrerad lista, indrag, tabeller, etc).",
     "Spara canvas → stäng app → öppna fil.",
     "Alla shape-fält (label, style, textAlignment, hasBullets, hasNumberedList, indentLevel, tableCells, position, size, rotation, etc) återskapas exakt."),
]


# ====================================================================
# Skapa Excel-arbetsboken
# ====================================================================
wb = Workbook()
ws = wb.active
ws.title = "Funktionsprotokoll v46"

# Header
headers = [
    "Nr",
    "Kategori / Meny",
    "Funktion",
    "Beskrivning",
    "Åtkomst (hur man kommer åt)",
    "Förväntad effekt",
    "Status (simulator)",
    "Status (iPhone)",
    "Anteckning / bugg-ID",
]
ws.append(headers)

# Title / metadata-rader ovan (vi flyttar header till rad 4, lägger titel ovan)
# Justera: vi rivar och börjar om med två titel-rader först.
ws.delete_rows(1)

ws["A1"] = "Visuali2e v46 — Funktions- och valideringsprotokoll"
ws["A1"].font = Font(name="Helvetica", size=18, bold=True, color="2E1A6B")
ws.merge_cells("A1:I1")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws["A2"] = ("Genererad 2026-05-21 efter v46-deploy. "
            "Använd Status-kolumnerna för att markera ✓ (OK), ✗ (fel) eller — (ej testat) per funktion. "
            "Testa varje funktion i simulator OCH på iPhone enligt CLAUDE.md regel #4.")
ws["A2"].font = Font(name="Helvetica", size=10, italic=True, color="555555")
ws.merge_cells("A2:I2")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 34

# Header på rad 4
ws.row_dimensions[3].height = 6  # liten luft-rad
ws.append([])  # rad 3 = tom luft
ws.append(headers)  # rad 4 = header

header_row = 4
header_font = Font(name="Helvetica", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E1A6B", end_color="2E1A6B", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="888888")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[header_row].height = 36

# Data-rader — alternerande fill per kategori
data_start = header_row + 1
category_color_map = {}
palette = [
    "F4F0FB",  # ljust lila
    "EEF6FC",  # ljust blå
    "FFF4E6",  # ljust persika
    "EFFAEF",  # ljust grön
    "FCF1F4",  # ljust rosa
    "F5F5F5",  # ljus grå
    "FFF7DF",  # ljus gul
    "F0EBFA",  # lavendel
    "E8F3FA",  # himmel
    "FAEFE6",  # sand
]
palette_idx = 0


def col_for(category: str) -> str:
    global palette_idx
    if category not in category_color_map:
        category_color_map[category] = palette[palette_idx % len(palette)]
        palette_idx += 1
    return category_color_map[category]


for i, (cat, name, desc, access, effect) in enumerate(FUNCTIONS, start=1):
    row_idx = data_start + (i - 1)
    fill_color = col_for(cat)
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    values = [i, cat, name, desc, access, effect, "", "", ""]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        cell.font = Font(name="Helvetica", size=10)
        cell.border = border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.font = Font(name="Helvetica", size=10, bold=True, color="2E1A6B")
        if col_idx == 2:
            cell.font = Font(name="Helvetica", size=10, bold=True, color="2E1A6B")
        if col_idx == 3:
            cell.font = Font(name="Helvetica", size=10, bold=True)
        if col_idx in (7, 8):
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Kolumnbredder (i ungefärliga Excel-units, tweakade för läsbarhet)
widths = {
    "A": 5,     # Nr
    "B": 22,    # Kategori
    "C": 28,    # Funktion
    "D": 52,    # Beskrivning
    "E": 48,    # Åtkomst
    "F": 48,    # Förväntad effekt
    "G": 16,    # Status sim
    "H": 16,    # Status iPhone
    "I": 26,    # Anteckning
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Frys översta rader (titel + header)
ws.freeze_panes = f"A{header_row + 1}"

# Sätt höjd på alla data-rader till "auto" via en rimlig default
for i in range(data_start, data_start + len(FUNCTIONS)):
    ws.row_dimensions[i].height = 50

# Spara
out_path = "/Users/kim/2e Mermaid Code/Visuali2e-v46-Funktionsprotokoll.xlsx"
wb.save(out_path)
print(f"Sparad: {out_path}")
print(f"Totalt antal funktioner: {len(FUNCTIONS)}")
print(f"Kategorier: {len(set(c for c, *_ in FUNCTIONS))}")
